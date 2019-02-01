from django.shortcuts import render
from rest_framework import viewsets , permissions , serializers
from url_filter.integrations.drf import DjangoFilterBackend
from .serializers import *
from API.permissions import *
from .models import *
from django.db.models import Q, F , Sum
from openpyxl import load_workbook,Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import PatternFill , Font
import string
from rest_framework.views import APIView
from rest_framework.renderers import JSONRenderer
from rest_framework.response import Response
from io import BytesIO
from rest_framework.views import APIView
from projects.models import ProjectPettyExpense , project
import datetime
from dateutil.relativedelta import relativedelta
import calendar
from django.http import HttpResponse
from reportlab import *
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import cm, mm
from reportlab.lib import colors , utils
from reportlab.platypus import Paragraph, Table, TableStyle, Image, Frame, Spacer, PageBreak, BaseDocTemplate, PageTemplate, SimpleDocTemplate, Flowable,ListItem,ListFlowable,NextPageTemplate
from PIL import Image
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet, TA_CENTER
from reportlab.graphics import barcode , renderPDF
from reportlab.graphics.shapes import *
from reportlab.graphics.barcode.qr import QrCodeWidget
from reportlab.lib.pagesizes import letter,A5,A4,A3,A2
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import ParagraphStyle,getSampleStyleSheet
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from reportlab.lib.colors import *
from reportlab.lib.units import inch, cm
from num2words import num2words
from django.core.mail import send_mail, EmailMessage
from django.db.models import CharField,FloatField, Value , Func
from clientRelationships.models import Contract
from openpyxl import load_workbook,Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import PatternFill , Font
# Create your views here.

class AccountViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated, isAdmin, )
    serializer_class = AccountSerializer
    queryset = Account.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['number','personal','contactPerson']

class InflowViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated, isAdmin, )
    serializer_class = InflowSerializer
    queryset = Inflow.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['amount']
    def get_queryset(self):
        if 'search' in self.request.GET:
            try:
                toRet = Inflow.objects.filter(amount__gte=int(self.request.GET['search']))
                return toRet
            except:
                return Inflow.objects.all()
        else:
            return Inflow.objects.all()

class CostCenterViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated, isAdmin, )
    serializer_class = CostCenterSerializer
    queryset = CostCenter.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['name']

class TransactionViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = TransactionSerializer
    # queryset = Transaction.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['amount','toAcc']
    def get_queryset(self):
        if 'filterBoth' in self.request.GET:
            toRet = Transaction.objects.filter(Q(fromAcc__number=self.request.GET['filterBoth'])|Q(toAcc__number=self.request.GET['filterBoth']))
            return toRet
        else:
            return Transaction.objects.all()

class ExpenseSheetViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = ExpenseSheetSerializer
    queryset = ExpenseSheet.objects.all()

class ExpenseHeadingViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = ExpenseHeadingSerializer
    queryset = ExpenseHeading.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['title']

class InvoiceViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = InvoiceSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['description', 'dated','sheet']
    def get_queryset(self):
        u = self.request.user
        return Invoice.objects.all()

class VendorProfileViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = VendorProfileSerializer
    queryset = VendorProfile.objects.all()
    # filter_backends = [DjangoFilterBackend]
    # filter_fields = ['service']

class VendorServiceViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = VendorServiceSerializer
    queryset = VendorService.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['vendorProfile']

class VendorInvoiceViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = VendorInvoiceSerializer
    queryset = VendorInvoice.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['vendorProfile']

class PurchaseOrderViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = PurchaseOrderSerializer
    queryset = PurchaseOrder.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['name','poNumber','isInvoice','status']

class PurchaseOrderQtyViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = PurchaseOrderQtySerializer
    queryset = PurchaseOrderQty.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['purchaseorder','product']

class OutBoundInvoiceViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = OutBoundInvoiceSerializer
    queryset = OutBoundInvoice.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['poNumber','status']

class OutBoundInvoiceQtyViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = OutBoundInvoiceQtySerializer
    queryset = OutBoundInvoiceQty.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['outBound','product']

class InventoryViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = InventorySerializer
    queryset = Inventory.objects.all()
    # filter_backends = [DjangoFilterBackend]
    # filter_fields = ['outBound','product']

class InventoryLogViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = InventoryLogSerializer
    queryset = InventoryLog.objects.all()
    # filter_backends = [DjangoFilterBackend]
    # filter_fields = ['outBound','product']

class UplodInflowDataAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.IsAuthenticated ,)
    def post(self , request , format = None):
        print 'entered','*******************'
        requestData = request.data
        print requestData
        tosend={}
        if 'exFile' in requestData:
            excelFile = request.FILES['exFile']
            if request.data['verified'] in ['true','1','yes',]:
                verified = True
            else:
                verified = False
            currency = request.data['currency']
            toAcc = Account.objects.get(pk=int(request.data['toAcc']))
            user = request.user
            print verified,type(verified),toAcc,currency,user
            wb = load_workbook(filename = BytesIO(excelFile.read()))
            for idx,ws in enumerate(wb.worksheets):
                wsTitle = ws.title
                print wsTitle,idx
                if idx==0:
                    storeData = []
                    print 'savingggggggg'
                    for i in range(2,ws.max_row+1):
                        try:
                            amount = int(ws['A'+str(i)].value) if ws['A'+str(i)].value else 0
                            referenceID = str(ws['B'+str(i)].value) if ws['B'+str(i)].value else None
                            dated = ws['C'+str(i)].value.date() if ws['C'+str(i)].value else None
                            description = str(ws['D'+str(i)].value) if ws['D'+str(i)].value else None
                            chequeNo = str(ws['E'+str(i)].value) if ws['E'+str(i)].value else None
                            mode = str(ws['F'+str(i)].value) if ws['F'+str(i)].value else 'cash'
                            mode = mode.lower()
                            gstCollected = float(ws['G'+str(i)].value) if ws['G'+str(i)].value else 0

                            infObj = Inflow(user=user,toAcc=toAcc,currency=currency,verified=verified,amount=amount,referenceID=referenceID,dated=dated,description=description,chequeNo=chequeNo,mode=mode,gstCollected=gstCollected)
                            storeData.append(infObj)
                        except:
                            print 'row number {0} in Excel - {1} is not created'.format(i,wsTitle)
                    Inflow.objects.bulk_create(storeData)
                    print 'total {0} Objects has been created'.format(len(storeData))
        return Response(tosend, status=status.HTTP_200_OK)

class GetExpenseDataAPI(APIView):
    permission_classes = (permissions.IsAuthenticated ,)
    def get(self , request , format = None):
        print 'entered','*******************'
        print request.GET
        accountsList = list(request.user.accountsManaging.all().values_list('pk',flat=True).distinct())
        print accountsList,'user accountssssss Listttttttt'
        tosend=[]
        expObj = ProjectPettyExpense.objects.filter(account__in=accountsList).values('project__pk','project__title').distinct()
        print expObj
        for i in expObj:
            expTotal = ProjectPettyExpense.objects.filter(project__id=int(i['project__pk']),account__in=accountsList).aggregate(tot=Sum('amount'))
            expTotal = expTotal['tot'] if expTotal['tot'] else 0
            print expTotal
            data = {'projectPk':i['project__pk'],'projectName':i['project__title'],'expTotal':expTotal}
            tosend.append(data)
        if 'limit' in request.GET:
            try:
                lmt = int(request.GET['limit'])
                tosend = tosend[0:lmt]
            except:
                pass
        return Response(tosend, status=status.HTTP_200_OK)

class ExpensesGraphDataAPI(APIView):
    permission_classes = (permissions.IsAuthenticated ,)
    def get(self , request , format = None):
        print 'entered','*******************'
        print request.GET

        tosend={'labels':[],'datasets':[]}
        allExpenses = ProjectPettyExpense.objects.all()
        projLists = allExpenses.values('project__pk','project__title').distinct()
        accountList = allExpenses.values('account__pk','account__title').distinct()
        for idx,i in enumerate(accountList):
            data = {'label':i['account__title'],'data':[]}
            for j in projLists:
                if idx==0:
                    tosend['labels'].append(j['project__title'])
                expTotal = allExpenses.filter(project__id=int(j['project__pk']),account__id=int(i['account__pk'])).aggregate(tot=Sum('amount'))
                expTotal = expTotal['tot'] if expTotal['tot'] else 0
                data['data'].append(expTotal)
            tosend['datasets'].append(data)

        return Response(tosend, status=status.HTTP_200_OK)

class MonthsExpensesDataAPI(APIView):
    permission_classes = (permissions.IsAuthenticated ,)
    def get(self , request , format = None):
        print 'entered','*******************'
        tosend={'labels':[],'datasets':[]}
        today = datetime.date.today()
        lstYear = today + relativedelta(years=-1,months=-1)
        print today , lstYear
        allExpenses = ProjectPettyExpense.objects.all()
        for i in range(13):
            lstYear += relativedelta(months=1)
            mth = lstYear.month
            yr = lstYear.year
            label = str(calendar.month_abbr[mth]) + '-' + str(yr)
            tosend['labels'].append(label)
            expTotal = allExpenses.filter(created__year=yr,created__month=mth).aggregate(tot=Sum('amount'))
            expTotal = expTotal['tot'] if expTotal['tot'] else 0
            tosend['datasets'].append(expTotal)

        return Response(tosend, status=status.HTTP_200_OK)

class DownloadExpenseSummaryAPI(APIView):
    permission_classes = (permissions.IsAuthenticated ,)
    def get(self , request , format = None):
        print 'entered','*******************'

        hdFont = Font(size=12,bold=True)
        alphaChars = list(string.ascii_uppercase)
        workbook = Workbook()
        Sheet1 = workbook.active
        Sheet1.title = 'Expense Summary'
        hd = ["Project", "Budget",'Expense']
        hdWidth = [10,10,10]
        Sheet1.append(hd)
        projectsList = project.objects.filter(projectClosed=False)
        print projectsList.count()
        for i in projectsList:
            data = [i.title,i.budget]
            try:
                expTotal = ProjectPettyExpense.objects.filter(project=i).aggregate(tot=Sum('amount'))
                expTotal = expTotal['tot'] if expTotal['tot'] else 0
            except:
                expTotal = 0
            data.append(expTotal)
            Sheet1.append(data)
            for idx,j in enumerate(data):
                if (len(str(j))+5) > hdWidth[idx]:
                    hdWidth[idx] = len(str(j)) + 5
        for idx,i in enumerate(hd):
            cl = str(alphaChars[idx])+'1'
            Sheet1[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
            Sheet1[cl].font = hdFont
            Sheet1.column_dimensions[str(alphaChars[idx])].width = hdWidth[idx]

        allExpenses = ProjectPettyExpense.objects.all()
        projLists = allExpenses.values('project__pk','project__title').distinct()
        for i in projLists:
            Sheet = workbook.create_sheet(i['project__title'])
            hd = ['Title','Amount','Account No.','User Name','Description']
            hdWidth = [10,10,10,10,10]
            Sheet.append(hd)

            ptObjs = allExpenses.filter(project__id=int(i['project__pk']))
            for j in ptObjs:
                uName = j.createdUser.first_name
                if j.createdUser.last_name:
                    uName += ' ' + j.createdUser.last_name
                data = [j.heading.title,j.amount,j.account.number,uName,j.description]
                Sheet.append(data)
                for idx,k in enumerate(data):
                    if (len(str(k))+5) > hdWidth[idx]:
                        hdWidth[idx] = len(str(k)) + 5
            for idx,j in enumerate(hd):
                cl = str(alphaChars[idx])+'1'
                Sheet[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
                Sheet[cl].font = hdFont
                Sheet.column_dimensions[str(alphaChars[idx])].width = hdWidth[idx]

        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=ExpenseSummary.xlsx'
        return response


def grn(response , project , purchaselist , request):
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(response,pagesize=letter, topMargin=0.2*cm,leftMargin=0.1*cm,rightMargin=0.1*cm)
    doc.request = request
    elements = []

    p1 = Paragraph("<para alignment='center'fontSize=15  ><b> Goods Received Note </b></para>",styles['Normal'])
    elements.append(p1)
    elements.append(Spacer(1, 10))
    try:
        address = project.address.replace('\n', '<br />')
    except:
        address = project.address
    addrdetails = Paragraph("""
    <para >
    <b>%s</b><br/>
    %s <br/>
    </para>
    """ %(project.name ,address),styles['Normal'])
    td=[[addrdetails]]
    t=Table(td,colWidths=[4*inch])
    t.hAlign = 'LEFT'
    elements.append(t)
    elements.append(Spacer(1,10))
    # details = Paragraph("""
    # <para>
    # <b>PO Number - </b>%s<br/>
    # </para>
    # """ %(project.poNumber),styles['Normal'])
    # td=[[details]]
    # t=Table(td,colWidths=[4*inch])
    # t.hAlign = 'LEFT'
    # elements.append(t)
    p9_01 =Paragraph("<para fontSize=8> <b>Sl. no</b></para>",styles['Normal'])
    p9_02 =Paragraph("<para fontSize=8><b>Product</b></para>",styles['Normal'])
    p9_03 =Paragraph("<para fontSize=8><b>Quantity</b></para>",styles['Normal'])
    data2=[[p9_01,p9_02,p9_03]]
    id=0
    for i in purchaselist:
        id+=1
        product = i.product
        quanty = i.receivedQty
        p10_01 =Paragraph("<para fontSize=8>{0}</para>".format(id),styles['Normal'])
        p10_02 =Paragraph("<para fontSize=8>{0}</para>".format(product),styles['Normal'])
        p10_03 =Paragraph("<para fontSize=8>{0}</para>".format(quanty),styles['Normal'])
        data2.append([p10_01,p10_02,p10_03])
    t3=Table(data2,colWidths=[0.5*inch , 3*inch , 1*inch])
    t3.hAlign = 'LEFT'
    t3.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'LEFT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t3)
    doc.build(elements)


class GrnAPIView(APIView):
    def get(self , request , format = None):
        project = PurchaseOrder.objects.get(pk = request.GET['value'])
        purchaselist = PurchaseOrderQty.objects.filter(purchaseorder = request.GET['value'])
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="Grndownload.pdf"'
        grn(response , project , purchaselist , request)
        return response

settingsFields = application.objects.get(name='app.clientRelationships').settings.all()
def invoice(response , inv , invdetails , typ, request):
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(response,pagesize=letter, topMargin=0.2*cm,leftMargin=0.1*cm,rightMargin=0.1*cm)
    doc.request = request
    elements = []
    print settingsFields.get(name='companyName').value,'aaaaaaaaaa'
    headerDetails = Paragraph("""
    <para align='center'>
    <font size ='8'>
    <b>%s</b><br/>
    %s<br/>
    Phone : 000000000<br/>
    </font>
    </para>
    """ %(settingsFields.get(name='companyName').value,settingsFields.get(name='companyAddress').value),styles['Normal'])
    tdheader=[['',headerDetails,'']]
    headerTitle = Paragraph("""
    <para align='center'>
    <font size ='12'>
    <b> Tax Invoice</b></font>
    </para>
    """ %(),styles['Normal'])
    tdheader+=[['',headerTitle,'']]
    t2=Table(tdheader,colWidths=(54.7*mm,100*mm,54.7*mm))
    t2.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black), ('LINEABOVE', (0,1), (-1,-1), 0.25, colors.black),]))
    elements.append(t2)
    detail31 = Paragraph("""
    <para align='center'>
    <b>Billing Address</b>
    </para>
    """ %(),styles['Normal'])
    detail32 = Paragraph("""
    <para align='center'>
    <b>Shipping Address</b>
    </para>
    """ %(),styles['Normal'])
    t2data=[detail31],[detail32]
    td2header=[t2data]
    t5=Table(td2header)
    t5.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t5)
    detail41 = Paragraph("""
    <para>
    Name :
    </para>
    """ %(),styles['Normal'])
    detail42 = Paragraph("""
    <para >
    %s
    </para>
    """ %(inv.name),styles['Normal'])
    detail43 = Paragraph("""
    <para >
    Name :
    </para>
    """ %(),styles['Normal'])
    detail44 = Paragraph("""
    <para >
    %s
    </para>
    """ %(inv.name),styles['Normal'])
    t4data=[detail41],[detail42],[detail43],[detail44]
    td4header=[t4data]
    t6=Table(td4header)
    t6.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t6)
    try:
        billaddr = inv.address.replace('\n', '<br />')
    except:
        billaddr = inv.address
    try:
        shipaddr = inv.address.replace('\n', '<br />')
    except:
        shipaddr = inv.address

    detail51 = Paragraph("""
    <para>
    Address : <br/> %s <br/>
    %s
    </para>
    """ %(billaddr,inv.pincode),styles['Normal'])
    detail52 = Paragraph("""
    <para>
    Address : <br/> %s <br/>
    %s
    </para>
    """ %(shipaddr,inv.pincode),styles['Normal'])
    t5data=[detail51],[detail52]
    td5header=[t5data]
    detail61 = Paragraph("""
    <para>
    GSTIN : %s
    </para>
    """ %(inv.gstIn),styles['Normal'])
    detail62 = Paragraph("""
    <para>
    GSTIN : %s
    </para>
    """ %(inv.gstIn),styles['Normal'])
    t6data=[detail61],[detail62]
    td5header+=[t6data]
    t7=Table(td5header)
    t7.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t7)
    data2=[]
    s01 =Paragraph("<para fontSize=8>S.No </para>",styles['Normal'])
    s02 =Paragraph("<para fontSize=8>Product </para>",styles['Normal'])
    s03 =Paragraph("<para fontSize=8>Qty </para>",styles['Normal'])
    s04 =Paragraph("<para fontSize=8>Rate </para>",styles['Normal'])
    s05 =Paragraph("<para fontSize=8>HSN/SAC Code</para>",styles['Normal'])
    s06 =Paragraph("<para fontSize=8>Taxable (%)  </para>",styles['Normal'])
    s07 =Paragraph("<para fontSize=8> Total </para>",styles['Normal'])
    data2 += [[s01,s02,s03,s04,s05,s06,s07]]
    id = 0
    grandtot= 0
    for i in invdetails:
        id+=1
        grandtot +=i.total
        print grandtot
        s21 =Paragraph("<para fontSize=8>{0} </para>".format(id),styles['Normal'])
        s22 =Paragraph("<para fontSize=8>{0} </para>".format(i.product),styles['Normal'])
        s23 =Paragraph("<para fontSize=8 alignment='center'>{0} </para>".format(i.qty),styles['Normal'])
        s24 =Paragraph("<para fontSize=8  alignment='right'>{:,} </para>".format(round(i.price,2)),styles['Normal'])
        try:
            hsn = i.hsn.code
        except:
             hsn = i.hsn
        s25 =Paragraph("<para fontSize=8 alignment='right'> {0}</para>".format(hsn),styles['Normal'])
        s26 =Paragraph("<para fontSize=8 alignment='right'>{0} </para>".format(i.tax),styles['Normal'])
        s27 =Paragraph("<para fontSize=8 alignment='right'>{:,} </para>".format(round(i.total,2)),styles['Normal'])
        data2.append([s21,s22,s23,s24,s25,s26,s27])
    s21 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
    s22 =Paragraph("<para fontSize=8></para>",styles['Normal'])
    s23 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
    s24 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
    s25 =Paragraph("<para fontSize=8  alignment='right'><b>Total</b></para>",styles['Normal'])
    s26 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
    s27 =Paragraph("<para fontSize=8 alignment='right'><b>{:,}</b> </para>".format(round(grandtot,2)),styles['Normal'])
    data2.append([s21,s22,s23,s24,s25,s26,s27])
    t9=Table(data2)
    t9.hAlign = 'LEFT'
    t9.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t9)
    gtotalText = num2words(int(grandtot), to='cardinal', lang='en_IN')
    print gtotalText
    s41 =Paragraph("<para fontSize=8> Rupees {0} </para>".format(gtotalText),styles['Normal'])
    datawords =[[s41]]
    t10=Table(datawords)
    t10.hAlign = 'LEFT'
    t10.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t10)
    if typ=='inbond':
        try:
            invtrms = inv.invoiceTerms.replace('\n', '<br />')
        except:
            invtrms = inv.invoiceTerms
        dataFooter = []
        s51 =Paragraph("<para fontSize=8>Payment Terms : {0} </para>".format(invtrms),styles['Normal'])
        s52 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
        s53 =Paragraph("<para fontSize=6 alignment='center'> Certified that the particulars given above are true and correct <br/></para><para fontSize=10> {0}</para>".format(settingsFields.get(name='companyName').value),styles['Normal'])
        dataFooter =[[s51,s52,s53]]
        s61 =Paragraph("<para fontSize=8>{0}</para>".format(settingsFields.get(name='bankDetails').value),styles['Normal'])
        s62 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
        s63 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
        dataFooter +=[[s61,s62,s63]]
        t11=Table(dataFooter,colWidths=(80*mm,49.5*mm,80*mm))
        # t11.hAlign = 'LEFT'
        t11.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black), ('LINEABOVE', (1,1), (-1,-1), 0.50, colors.white),]))
        elements.append(t11)
    else:
        dataFooter = []
        # s51 =Paragraph("<para fontSize=8>Payment Terms : {0} </para>".format(invtrms),styles['Normal'])
        # s52 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
        # s53 =Paragraph("<para fontSize=6 alignment='center'> Certified that the particulars given above are true and correct <br/></para><para fontSize=10> {0}</para>".format(settingsFields.get(name='companyName').value),styles['Normal'])
        # dataFooter =[[s51,s52,s53]]
        s61 =Paragraph("<para fontSize=8>{0}</para>".format(settingsFields.get(name='bankDetails').value),styles['Normal'])
        s62 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
        s63 =Paragraph("<para fontSize=6 alignment='center'> Certified that the particulars given above are true and correct <br/></para><para fontSize=10> {0}</para>".format(settingsFields.get(name='companyName').value),styles['Normal'])
        dataFooter +=[[s61,s62,s63]]
        t11=Table(dataFooter,colWidths=(80*mm,49.5*mm,80*mm))
        # t11.hAlign = 'LEFT'
        t11.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black), ('LINEABOVE', (1,1), (-1,-1), 0.25, colors.white),]))
        elements.append(t11)
    doc.build(elements)

class InvoiceAPIView(APIView):
    def get(self , request , format = None):
        typ = request.GET['typ']
        if typ =='inbond':
            inv = PurchaseOrder.objects.get(pk = request.GET['value'])
            invDetails = PurchaseOrderQty.objects.filter(purchaseorder = request.GET['value'])
        else:
            inv = OutBoundInvoice.objects.get(pk = request.GET['value'])
            invDetails = OutBoundInvoiceQty.objects.filter(outBound = request.GET['value'])
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="Invoicedownload.pdf"'
        invoice(response , inv , invDetails , typ, request)
        return response

class SendInvoiceAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def post(self, request, format=None):
        contactData=[]
        print request.data,'sssssssss'
        typ = request.data['typ']
        if typ=='inbond':
            inv = PurchaseOrder.objects.get(pk = request.data['value'])
            invDetails = PurchaseOrderQty.objects.filter(purchaseorder = request.data['value'])
        else:
            inv = OutBoundInvoice.objects.get(pk = request.data['value'])
            invDetails = OutBoundInvoiceQty.objects.filter(outBound = request.data['value'])
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="Invoicedownload.pdf"'
        invoice(response , inv , invDetails , typ, request)
        f = open(os.path.join(globalSettings.BASE_DIR, 'media_root/Invoicedownload.pdf'), 'wb')
        f.write(response.content)
        f.close()
        email_subject = "Invoice"
        msgBody = "Hi "  + inv.name + ",\n\n\t\t Find the attachment for the Invoice Details"
        email = request.data['email']
        contactData = tuple(email.split(','))
        print contactData
        # contactData.append(str(email))
        msg = EmailMessage(email_subject, msgBody,  to= contactData )
        a = str(f).split('media_root/')[1]
        b = str(a).split(', mode')[0]
        c = str(b).split("'")[0]
        msg.attach_file(os.path.join(globalSettings.MEDIA_ROOT,str(c)))
        msg.send()
        return Response({}, status = status.HTTP_200_OK)


# class SendInvoicingInvoiceAPIView(APIView):
#     renderer_classes = (JSONRenderer,)
#     def post(self, request, format=None):
#         contactData=[]
#         print request.data,'sssssssss'
#         inv = PurchaseOrder.objects.get(pk = request.data['value'])
#         invDetails = PurchaseOrderQty.objects.filter(purchaseorder = request.data['value'])
#         response = HttpResponse(content_type='application/pdf')
#         response['Content-Disposition'] = 'attachment;filename="Invoicedownload.pdf"'
#         invoice(response , inv , invDetails , request)
#         f = open(os.path.join(globalSettings.BASE_DIR, 'media_root/Invoicedownload.pdf'), 'wb')
#         f.write(response.content)
#         f.close()
#         email_subject = "Invoice"
#         msgBody = "Hi "  + inv.name + ",\n\n\t\t Find the attachment for the Invoice Details"
#         email = request.data['email']
#         contactData = tuple(email.split(','))
#         print contactData
#         # contactData.append(str(email))
#         msg = EmailMessage(email_subject, msgBody,  to= contactData )
#         a = str(f).split('media_root/')[1]
#         b = str(a).split(', mode')[0]
#         c = str(b).split("'")[0]
#         msg.attach_file(os.path.join(globalSettings.MEDIA_ROOT,str(c)))
#         msg.send()
#         return Response({}, status = status.HTTP_200_OK)

class AverageAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        inboundobj = Inflow.objects.all()
        invobj = OutBoundInvoiceQty.objects.all()
        crmobj = Contract.objects.all()
        inboundtotal = 0
        invgst = 0
        inboundtotal = inboundobj.aggregate(inboundtotal=Sum(F('amount') ,output_field=FloatField())).get('inboundtotal',0)
        inboundgst = inboundobj.aggregate(inboundgst=Sum(F('gstCollected') ,output_field=FloatField())).get('inboundgst',0)
        invtotal = 0
        invgst = 0
        invtotal = invobj.aggregate(invtotal=Sum(F('total') ,output_field=FloatField())).get('invtotal',0)
        invgst = invobj.aggregate(invgst=Sum(((F('total')*F('tax'))/100) ,output_field=FloatField())).get('invgst',0)
        crmtotal = 0
        crmgst = 0
        crmtotal = crmobj.aggregate(crmtotal=Sum(F('grandTotal') ,output_field=FloatField())).get('crmtotal',0)
        crmgst = crmobj.aggregate(crmgst=Sum(F('totalTax') ,output_field=FloatField())).get('crmgst',0)
        overallTotal = inboundtotal + invtotal + crmtotal
        overallGst = inboundgst + invgst + crmgst
        return Response({'inboundtotal':inboundtotal,"inboundgst":inboundgst,'invtotal':invtotal,'invgst':invgst,'crmtotal':crmtotal,'crmgst':crmgst,'overallTotal':overallTotal,'overallGst':overallGst}, status = status.HTTP_200_OK)

class InvoiceSheetAPIView(APIView):
    def get(self, request, format=None):
        workbook = Workbook()
        Sheet1 = workbook.active
        hdFont = Font(size=12,bold=True)
        alphaChars = list(string.ascii_uppercase)
        Sheet1.title = 'External Invoices'
        hd = ["Bank", "Reference Id",'Amount','Description','GST']
        hdWidth = [10,10,10,10,10,10,10]
        Sheet1.append(hd)
        inboundobj = Inflow.objects.all()
        for i in inboundobj:
            Sheet1.append([i.toAcc.title,i.referenceID,i.amount,i.description,i.gstCollected])
        if Sheet1.max_column <= len(alphaChars):
            for character in alphaChars[0:Sheet1.max_column]:
                Sheet1.column_dimensions[character].width = 20
        for idx,i in enumerate(hd):
            cl = str(alphaChars[idx])+'1'
            Sheet1[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
            Sheet1[cl].font = hdFont
        Sheet2 = workbook.create_sheet('CRM Invoices')
        hd = ["Id","Dated", " Total Value",'Grand Total','Total Tax']
        Sheet2.append(hd)
        crmobj = Contract.objects.all()
        for c in crmobj:
            try:
                dated = str(c.recievedDate).split(' ')[0]
                if dated == "None":
                    dated = ''
            except:
                dated = ''
            Sheet2.append([c.pk,dated,c.value,c.grandTotal,c.totalTax])
        if Sheet2.max_column <= len(alphaChars):
            for character in alphaChars[0:Sheet2.max_column]:
                Sheet2.column_dimensions[character].width = 20
        for idx,i in enumerate(hd):
            cl = str(alphaChars[idx])+'1'
            Sheet2[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
            Sheet2[cl].font = hdFont
        Sheet3 = workbook.create_sheet('Invoices')
        hd = ["Id", " Total",'Total Tax','Grand Total']
        Sheet3.append(hd)
        invobj = OutBoundInvoiceQty.objects.all()
        for iv in invobj:
            tax = ((iv.total*iv.tax)/100)
            tot = iv.total - tax
            Sheet3.append([iv.pk,tot,tax,iv.total])
        if Sheet3.max_column <= len(alphaChars):
            for character in alphaChars[0:Sheet3.max_column]:
                Sheet3.column_dimensions[character].width = 20
        for idx,i in enumerate(hd):
            cl = str(alphaChars[idx])+'1'
            Sheet3[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
            Sheet3[cl].font = hdFont
        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=stockConsumed.xlsx'
        return response
