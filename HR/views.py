from django.contrib.auth.models import User , Group
from django.shortcuts import render, redirect , get_object_or_404
from django.contrib.auth import authenticate , login , logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.urlresolvers import reverse
from django.template import RequestContext
from django.conf import settings as globalSettings
from django.core.exceptions import ObjectDoesNotExist , SuspiciousOperation
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from rest_framework.renderers import JSONRenderer
# Related to the REST Framework
from rest_framework import viewsets , permissions , serializers
from rest_framework.exceptions import *
from url_filter.integrations.drf import DjangoFilterBackend
from .serializers import *
from API.permissions import *
from ERP.models import application, permission , module
from ERP.views import getApps, getModules
from django.db.models import Q
from django.http import JsonResponse
from rest_framework.response import Response
import random, string
import requests
from django.utils import timezone
from rest_framework.views import APIView
from ecommerce.models import GenericImage
from django.template.loader import render_to_string, get_template
from django.core.mail import send_mail, EmailMessage

from openpyxl import load_workbook
from io import BytesIO

from rest_framework.authentication import SessionAuthentication, BasicAuthentication

class CsrfExemptSessionAuthentication(SessionAuthentication):

    def enforce_csrf(self, request):
        return  # To not perform the csrf check previously happening


def documentView(request):
    docID = None
    if request.method == 'POST':
        templt = 'documentVerify.external.showDetails.html'
        docID = request.POST['id']
        passKey = request.POST['passkey']

    elif request.method == 'GET':
        if 'id' in request.GET:
            docID = request.GET['id']
            passKey = request.GET['passkey']
            templt = 'documentVerify.external.showDetails.html'
        else:
            templt = 'documentVerify.external.getPassKey.html'

    if docID is not None:
        if len(docID)<5:
            raise ObjectDoesNotExist("Document ID not correct")
        doc = get_object_or_404(Document , pk = int(docID[4:]), passKey = passKey)
        templt = 'documentVerify.external.showDetails.html'
        eml = doc.email
        prts = eml.split('@')
        eml = prts[0][0:4]+ "*******@" + prts[1]
        data = {
            "id":doc.pk,
            "issuedTo" : doc.issuedTo,
            "issuedBy" : doc.issuedBy,
            "created" : doc.created,
            "description" : doc.description,
            "email": eml
        }

    else:
        data = {}

    return render(request , templt, data)

def generateOTPCode():
    length = 4
    chars = string.digits
    rnd = random.SystemRandom()
    return ''.join(rnd.choice(chars) for i in range(length))

def tokenAuthentication(request):

    ak = get_object_or_404(accountsKey, activation_key=request.GET['key'] , keyType='hashed')
    #check if the activation key has expired, if it hase then render confirm_expired.html
    if ak.key_expires < timezone.now():
        raise SuspiciousOperation('Expired')
    #if the key hasn't expired save user and set him as active and render some template to confirm activation
    user = ak.user
    user.is_active = True
    user.save()
    user.accessibleApps.all().delete()
    for a in globalSettings.DEFAULT_APPS_ON_REGISTER:
        app = application.objects.get(name = a)
        p = permission.objects.create(app =  app, user = user , givenBy = User.objects.get(pk=1))
    login(request , user)
    authStatus = {'status' : 'success' , 'message' : 'Account actived, please login.' }
    return render(request , globalSettings.LOGIN_TEMPLATE , {'authStatus' : authStatus ,'useCDN' : globalSettings.USE_CDN})

@csrf_exempt
def generateOTP(request):
    print request.POST ,'*****************'
    print request.POST["id"],'kkkkkkkkllllllllllllkkkkkkkkkkk'

    key_expires = timezone.now() + datetime.timedelta(2)
    otp = generateOTPCode()
    print 'uuuuuuuuuuuuuuuuuuuuuuuuuuuuu'
    user = get_object_or_404(User, username = request.POST['id'])
    print user,type(user),'uuuuuuuuuuuuuuuuuuuuuuuuuuuuu'
    ak = accountsKey(user= user, activation_key= otp,
        key_expires=key_expires , keyType = 'otp')
    ak.save()
    print ak.activation_key
    # send a SMS with the OTP
    url = globalSettings.SMS_API_PREFIX + 'number=%s&message=%s'%(request.POST['id'] , 'Dear Customer,\nPlease use OTP : %s to verify your mobile number' %(otp))
    requests.get(url)
    return JsonResponse({} ,status =200 )

@csrf_exempt
def loginView(request):

    # print request.META['HTTP_USER_AGENT']
    print 'cameeeeeeeeeeeeeeeeeeeeeeeeeeeeeeee',request

    backgroundImage = globalSettings.LOGIN_PAGE_IMAGE
    genericImg = GenericImage.objects.all()
    try:
        if len(genericImg)>0:
            if genericImg[0].backgroundImage:
                backgroundImage = genericImg[0].backgroundImage.url
    except:
        print 'no login imageeee'

    if globalSettings.LOGIN_URL != 'login':
        return redirect(reverse(globalSettings.LOGIN_URL))
    authStatus = {'status' : 'default' , 'message' : '' }
    statusCode = 200
    if request.user.is_authenticated():
        if request.GET:
            return redirect(request.GET['next'])
        else:
            return redirect(reverse(globalSettings.LOGIN_REDIRECT))
    if request.method == 'POST':
    	usernameOrEmail = request.POST['username']
        print usernameOrEmail ,'usernameOrEmail'
        otpMode = False
        if 'otp' in request.POST:
            print "otp"
            otp = request.POST['otp']
            otpMode = True
        else:
            password = request.POST['password']
        if '@' in usernameOrEmail and '.' in usernameOrEmail:
            try:
                u = User.objects.get(email = usernameOrEmail)
                username = u.username
            except:
                statusCode = 404
                username = usernameOrEmail
        else:
            username = usernameOrEmail
            try:
                u = User.objects.get(username = username)
            except:
                statusCode = 404
        if not otpMode:
            user = authenticate(username = username , password = password)
        else:
            print "OTP Mode"
            ak = None
            try:
                aks = accountsKey.objects.filter(activation_key=otp , keyType='otp')
                ak = aks[len(aks)-1]
                print "Aks", aks,ak
            except:
                pass
            print ak
            if ak is not None:
                #check if the activation key has expired, if it has then render confirm_expired.html
                if ak.key_expires > timezone.now():
                    user = ak.user
                    user.backend = 'django.contrib.auth.backends.ModelBackend'
                else:
                    user = None
            else:
                authStatus = {'status' : 'danger' , 'message' : 'Incorrect OTP'}
                statusCode = 401

    	if user is not None:
            login(request , user)
            if request.GET and 'next' in request.GET:
                return redirect(request.GET['next'])
            else:
                if 'mode' in request.GET and request.GET['mode'] == 'api':
                    return JsonResponse({} , status = 200)
                else:
                    return redirect(reverse(globalSettings.LOGIN_REDIRECT))
        else:
            if statusCode == 200 and not u.is_active:
                authStatus = {'status' : 'warning' , 'message' : 'Your account is not active.'}
                statusCode = 423
            else:
                authStatus = {'status' : 'danger' , 'message' : 'Incorrect username or password.'}
                statusCode = 401
    if 'mode' in request.GET and request.GET['mode'] == 'api':
        return JsonResponse(authStatus , status = statusCode)


    # return render(request , globalSettings.LOGIN_TEMPLATE , {'authStatus' : authStatus ,'useCDN' : globalSettings.USE_CDN , 'backgroundImage': backgroundImage , 'icon_logo': globalSettings.ICON_LOGO ,"brandLogo" : globalSettings.BRAND_LOGO , "brandLogoInverted": globalSettings.BRAND_LOGO_INVERT}, status=statusCode)

    if globalSettings.LITE_REGISTRATION:
        return render(request,"login.html" , {'authStatus' : authStatus ,'useCDN' : globalSettings.USE_CDN , 'backgroundImage': globalSettings.LOGIN_PAGE_IMAGE , 'icon_logo': globalSettings.ICON_LOGO ,"brandLogo" : globalSettings.BRAND_LOGO , "brandLogoInverted": globalSettings.BRAND_LOGO_INVERT}, status=statusCode)
    else:
        return render(request,"loginCopy.html" , {'authStatus' : authStatus ,'useCDN' : globalSettings.USE_CDN , 'backgroundImage': globalSettings.LOGIN_PAGE_IMAGE , 'icon_logo': globalSettings.ICON_LOGO ,"brandLogo" : globalSettings.BRAND_LOGO , "brandLogoInverted": globalSettings.BRAND_LOGO_INVERT}, status=statusCode)


def registerView(request):
    print 'registerrrrrrrrrrrrrrrrrrrrrrrrrrr',request.POST
    if globalSettings.REGISTER_URL != 'register':
        return redirect(reverse(globalSettings.REGISTER_URL))
    msg = {'status' : 'default' , 'message' : '' }
    if request.method == 'POST':
    	name = request.POST['name']
    	email = request.POST['email']
    	password = request.POST['password']
        if User.objects.filter(email = email).exists():
            msg = {'status' : 'danger' , 'message' : 'Email ID already exists' }
            print msg,'emaillllllllll'
        else:
            user = User.objects.create(username = email.replace('@' , '').replace('.' ,''))
            user.first_name = name
            user.email = email
            user.set_password(password)
            user.save()
            user = authenticate(username = email.replace('@' , '').replace('.' ,'') , password = password)
            login(request , user)
            if request.GET:
                return redirect(request.GET['next'])
            else:
                return redirect(globalSettings.LOGIN_REDIRECT)
    return render(request , 'register.simple.html' , {'msg' : msg})


def logoutView(request):
    logout(request)
    return redirect(globalSettings.LOGOUT_REDIRECT)

def root(request):
    return redirect(globalSettings.ROOT_APP)


@login_required(login_url = globalSettings.LOGIN_URL)
def home(request):
    u = request.user
    if u.is_superuser:
        apps = application.objects.all()
        modules = module.objects.filter(~Q(name='public'))
    else:
        apps = getApps(u)
        modules = getModules(u)

    defaultRoute = 'admin'


    if globalSettings.SHOW_COMMON_APPS:
        showCommonApps = 'true'
    else:
        showCommonApps = 'false'

    apps = apps.filter(~Q(name__startswith='configure.' )).filter(~Q(name='app.users')).filter(~Q(name__endswith='.public'))
    return render(request , 'ngBase.html' , {'wampServer' : globalSettings.WAMP_SERVER, 'appsWithJs' : apps.filter(haveJs=True) \
    ,'appsWithCss' : apps.filter(haveCss=True) , 'modules' : modules , 'useCDN' : globalSettings.USE_CDN , 'BRAND_LOGO' : globalSettings.BRAND_LOGO \
    ,'BRAND_NAME' :  globalSettings.BRAND_NAME, 'serviceName' : globalSettings.SERVICE_NAME , 'defaultRoute' : defaultRoute , 'showCommonApps' : showCommonApps})

class userProfileViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny,)
    serializer_class = userProfileSerializer
    queryset = profile.objects.all()

class userProfileAdminModeViewSet(viewsets.ModelViewSet):
    permission_classes = (isAdmin ,)
    serializer_class = userProfileAdminModeSerializer
    queryset = profile.objects.all()

class userDesignationViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    queryset = designation.objects.all()
    serializer_class = userDesignationSerializer

class userAdminViewSet(viewsets.ModelViewSet):
    permission_classes = (isAdmin ,)
    queryset = User.objects.all()
    serializer_class = userAdminSerializer

class UserViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated ,)
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['username']
    serializer_class = userSerializer
    def get_queryset(self):
        if 'mode' in self.request.GET:
            if self.request.GET['mode']=="mySelf":
                if self.request.user.is_authenticated:
                    return User.objects.filter(username = self.request.user.username)
                else:
                    raise PermissionDenied()
            else :
                return User.objects.all().order_by('-date_joined')
        else:
            return User.objects.all().order_by('-date_joined')

class UserSearchViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated ,)
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['username']
    serializer_class = userSearchSerializer
    queryset = User.objects.all()
    def get_queryset(self):
        if 'mode' in self.request.GET:
            if self.request.GET['mode']=="mySelf":
                if self.request.user.is_authenticated:
                    return User.objects.filter(username = self.request.user.username)
                else:
                    raise PermissionDenied()
            else :
                return User.objects.all().order_by('-date_joined')
        else:
            return User.objects.all().order_by('-date_joined')

class GroupViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    queryset = Group.objects.all()
    serializer_class = groupSerializer

class rankViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    queryset = rank.objects.all()
    serializer_class = rankSerializer

class payrollViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    queryset = payroll.objects.all()
    serializer_class = payrollSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['user' ]


class SendActivatedStatus(APIView):
    renderer_classes = (JSONRenderer,)
    def post(self , request , format = None):
        print request.data

        msgBody = ['You have successfully registered' ]

        ctx = {
            'heading' : 'Welcome to Ecommerce',
            'recieverName' : str(request.data['name']),
            'message': msgBody,
            'linkUrl': 'sterlingselect.com',
            'linkText' : 'View Online',
            'sendersAddress' : 'sterlingselect',
            'sendersPhone' : '841101',
            'linkedinUrl' : 'https://www.linkedin.com/company/24tutors/',
            'fbUrl' : 'https://www.facebook.com/24tutorsIndia/',
            'twitterUrl' : 'twitter.com',
            'brandName' : globalSettings.BRAND_NAME,
        }
        emailAddr = []
        emailAddr.append(str(request.data['email']))
        email_body = get_template('app.HR.userActivated.html').render(ctx)
        email_subject = 'Registration Successfull!!!!!'
        email_cc = []
        email_bcc = []
        send_email(email_body,emailAddr,email_subject,email_cc,email_bcc,'html')
        return Response({}, status = status.HTTP_200_OK)


from django.core.mail import send_mail , EmailMessage
from django.core.mail import EmailMultiAlternatives
import sendgrid

from django.contrib.auth.tokens import PasswordResetTokenGenerator

token_generator = PasswordResetTokenGenerator()
print "token : " , token_generator.make_token(User.objects.get(pk = 1))

# path = reverse("account_reset_password_from_key", kwargs=dict(uidb36=user_pk_to_url_str(user),                                      key=temp_key))


class BulkUserCreationAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated , isAdmin)
    def post(self, request, format=None):
        print request.data,'aaaaaaaa'
        location = request.data["locationData"]
        url = location.split('/ERP')[0]
        wb = load_workbook(filename = BytesIO(request.FILES['xl'].read()))
        ws = wb.worksheets[0]
        row_count = ws.max_row+1
        column_count = ws.max_column

        for i in range(2, row_count):
            try:
                username = ws['A' + str(i)].value
            except:
                username =""
            try:
                email = ws['B' + str(i)].value
            except:
                email =""
            try:
                first_name = ws['C' + str(i)].value
            except:
                first_name =" "
            print first_name,'aaaaa'
            try:
                last_name = ws['D' + str(i)].value
                if last_name == None:
                    last_name = first_name
                else:
                    last_name = last_name
            except:
                last_name = first_name

            try:
                mobile = ws['E' + str(i)].value
            except:
                mobile =""
            try:
                designation = ws['F' + str(i)].value
                if designation == 'manager' or 'admin' or 'director':
                    is_staff = True
                else:
                    is_staff = False
            except:
                designation =""
            try:
                send = User(username=username, email= email, first_name=first_name, last_name=last_name,is_staff=is_staff)
                send.save()
            except:
                continue
            pobj = profile.objects.get(pk=send.profile.pk)
            pobj.email = email
            pobj.mobile = mobile
            pobj.details = {"username":username,"email":email,"first_name":first_name,"last_name":last_name,"designation":designation,"mobile":mobile,"GST":""}
            pobj.save()
            ctx = {
                'heading' : "Welcome to " + globalSettings.SITE_ADDRESS ,
                'link' : url + '/accounts/password/reset/',
                'recieverName' : first_name + ' ' + last_name,
                'brandName' : globalSettings.BRAND_NAME,
                'siteAddress' : globalSettings.SITE_ADDRESS
            }
            sendAddr = []
            sendAddr.append(str(email))
            email_body = get_template('app.user.resetPassword.html').render(ctx)
            email_subject = "Welcome to " + globalSettings.SITE_ADDRESS
            email_cc = []
            email_bcc = []
            send_email(email_body,sendAddr,email_subject,email_cc,email_bcc,'html')

        return Response(status = status.HTTP_200_OK)

@csrf_exempt
def socialMobileView(request):
    print request.POST,'ddddddddddddddddd'
    if request.POST['secretKey'] == globalSettings.MOBILE_SECRET_KEY:
        u = User.objects.filter(username = request.POST['username'])
        if len(u)>0:
            print 'already exist'
        else:
            print 'create new'
            u = User(username = request.POST['username'])
            u.email = request.POST['email']
            fname = request.POST['email'].split('@')[0]
            u.first_name = fname
            u.is_active = True
            u.set_password(request.POST['password'])
            u.save()
        loginView(request)
    else:
        raise PermissionDenied()
    return render(request,"ngEcommerce.html")
