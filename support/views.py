# -*- coding: utf-8 -*-

from django.contrib.auth.models import User , Group
from django.shortcuts import render, redirect,HttpResponse
from django.contrib.auth import authenticate , login , logout
from django.contrib.auth.decorators import login_required
from django.core.urlresolvers import reverse
from django.template import RequestContext
from django.conf import settings as globalSettings
# Related to the REST Framework
from rest_framework import viewsets , permissions , serializers
from rest_framework.exceptions import *
from url_filter.integrations.drf import DjangoFilterBackend
from .serializers import *
from API.permissions import *
from django.db.models import Q,F,Value,CharField
from allauth.account.adapter import DefaultAccountAdapter
from rest_framework.views import APIView
from rest_framework.renderers import JSONRenderer
import requests
from django.contrib.auth.hashers import make_password,check_password
from excel_response import ExcelResponse
# import libreERP.Checksum as Checksum
from django.views.decorators.csrf import csrf_exempt
import urllib
import os
from os import path
import datetime
from datetime import timedelta
from Crypto.Cipher import DES
import base64
import hashlib
from Crypto.Cipher import AES
from Crypto import Random
from django.db.models.functions import Concat
from ERP.models import service
from django.template.loader import render_to_string, get_template
from django.core.mail import send_mail, EmailMessage
from django.utils import timezone
import re
regex = re.compile('^HTTP_')

BLOCK_SIZE = 16
pad = lambda s: s + (BLOCK_SIZE - len(s) % BLOCK_SIZE) * chr(BLOCK_SIZE - len(s) % BLOCK_SIZE)
unpad = lambda s: s[:-ord(s[len(s) - 1:])]

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from dateutil.relativedelta import relativedelta
from django.db.models import Sum, Count, Avg
from .models import *

# Create your views here.

class CustomerProfileViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny,)
    serializer_class = CustomerProfileSerializer
    queryset = CustomerProfile.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['service']

class SupportChatViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny,)
    serializer_class = SupportChatSerializer
    # queryset = SupportChat.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['uid','user']
    exclude_fields = ['id']
    def get_queryset(self):
        supChatObj = SupportChat.objects.all()
        u = self.request.user
        if u.is_anonymous():
            supChatObj = supChatObj.filter(is_hidden = False)
        if 'visitorReq' in self.request.GET:
            supChatObj = supChatObj.filter(is_hidden = False)
        if 'user__isnull' in self.request.GET:
            return SupportChat.objects.filter(user__isnull=True)
        if 'uid' in self.request.GET:
            return supChatObj.filter(uid = self.request.GET['uid'])
        if 'date' in self.request.GET:
            date = datetime.datetime.strptime(self.request.GET['date'], '%Y-%m-%d').date()
            return supChatObj.filter(created__startswith = date)
        else:
            return SupportChat.objects.all()
        return SupportChat.objects.all()


class GetMyUser(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        print '****** entered', request.GET
        if 'getCompanyDetails' in request.GET:
            custP = CustomerProfile.objects.filter(pk=request.GET['getCompanyDetails'])
            objjj=list(CustomerProfile.objects.filter(pk=request.GET['getCompanyDetails']).values_list('name',flat=True))
            serviceObj=list(service.objects.filter(pk = custP[0].service.pk).values_list('contactPerson',flat=True))
            serviceName=list(service.objects.filter(pk = custP[0].service.pk).values_list('name',flat=True))
            servicePk=list(service.objects.filter(pk = custP[0].service.pk).values_list('pk',flat=True))


            return Response({'cDetails':serviceName,'contactP':serviceObj,'servicePk':servicePk}, status=status.HTTP_200_OK)
        if 'allAgents' in request.GET:
            print '%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%'
            allAgents = list(User.objects.exclude(pk=self.request.user.pk).values_list('pk',flat=True))
            print allAgents,type(allAgents)
            return Response({'allAgents':allAgents}, status=status.HTTP_200_OK)
        if 'getChatThread' in request.GET:
            uidsList = list(ChatThread.objects.filter(uid = request.GET['uid']).values_list('pk','company').distinct())
            return Response({'chatThread':uidsList}, status=status.HTTP_200_OK)

        if 'getMyUser' in request.GET:
            # uidsList = list(SupportChat.objects.filter(user = self.request.GET['user']).values_list('uid',flat=True).distinct())
            uidsList = list(ChatThread.objects.filter(user = self.request.GET['user'],status='started').values_list('uid',flat=True).distinct())
            print uidsList , 'distinct'
            toSend = []
            # chatThreadObjs = ChatThread.objects.filter(uid__in=uidsList)
            for i in uidsList:
                try:
                    data = Visitor.objects.get(uid=i)
                    dic = {'uid':data.uid,'name':data.name ,'email':data.email}
                except:
                    dic = {'uid':i,'name':'' ,'email':''}
                # print ChatThread.objects.get(uid=i).company.pk
                # dic['companyPk'] = ChatThread.objects.get(uid=i).company.pk
                # dic['chatThreadPk'] = ChatThread.objects.get(uid=i).pk
                dic['companyPk'] = ChatThread.objects.filter(uid=i)[0].company.pk
                dic['chatThreadPk'] = ChatThread.objects.filter(uid=i)[0].pk
                dic['servicePk'] = CustomerProfile.objects.filter(pk = dic['companyPk'])[0].service.pk
                print dic
                toSend.append(dic)
            return Response(toSend, status=status.HTTP_200_OK)
        if 'getDetail' in request.GET:
            # uidsList = list(SupportChat.objects.filter(user = self.request.GET['user']).values_list('uid',flat=True).distinct())
            toSend = []
            currentUser=User.objects.filter(pk=request.GET['pk'])
            print currentUser
            toSend.append(currentUser.values())
            return Response(toSend, status=status.HTTP_200_OK)
        if 'getNewUser' in request.GET:
            print 'getNewUser'
            time_threshold = datetime.datetime.now() - timedelta(hours=4)
            # results = Widget.objects.filter(created__lt=time_threshold)
            uidsList = list(ChatThread.objects.filter(user__isnull = True ,status='started' , created__gt = time_threshold).values_list('uid',flat=True).distinct())
            toSend = []
            for i in uidsList:
                try:
                    data = Visitor.objects.get(uid=i)
                    dic = {'uid':data.uid,'name':data.name ,'email':data.email}
                except:
                    dic = {'uid':i,'name':'' ,'email':''}
                dic['companyPk'] = ChatThread.objects.filter(uid=i)[0].company.pk
                dic['chatThreadPk'] = ChatThread.objects.filter(uid=i)[0].pk
                dic['servicePk'] = CustomerProfile.objects.filter(pk = dic['companyPk'])[0].service.pk
                # print dic['me']
                # print Support.objects.filter(uid = i).count() , 'CCCCCCCCCCCCCCCCCCCC'
                toSend.append(dic)
            print toSend , 'FFFFFFFFFFFFFFFF'
            return Response(toSend, status=status.HTTP_200_OK)
        if 'getNewComp' in request.GET:
            userpk=request.GET['getNewComp']
            myServices = service.objects.filter(advisors = userpk)
            companies = [];
            for i in myServices:
                company = CustomerProfile.objects.filter(service = i)
                companies.append(company[0].pk)
            print companies,'*************************8'
            return Response(companies, status=status.HTTP_200_OK)
def createExcel(data):
    wb = Workbook()

    # dest_filename = 'empty_book.xlsx'
    ws1 = wb.active
    ws1.title = "UID Wise"
    heading = ['UID','Email','Company','User Message','Agent Message',]
    heading_font = Font(bold=True, size=14)
    ws1.append(heading)
    print ws1.column_dimensions,'widthhhhhhhhhhhhhhhhhhhh'
    for idx,cell in enumerate(ws1["1:1"]):
        cell.font = heading_font
        ws1.column_dimensions[get_column_letter(idx+1)].width = 25 if idx<3 else 50
    uidList = []

    for row in data:
        dt = 'MEDIA ('+row['attachmentType'] +' )' if row['attachmentType'] else row['message']
        if row['uid'] in uidList:
            if row['sentByAgent']:
                ws1.append(['','','','',dt])
            else:
                ws1.append(['','','',dt,''])
        else:
            if len(uidList)>0:
                ws1.append(['','','','','',''])
                ws1.append(['','','','','',''])
                ws1.append(['','','','','',''])
            uidList.append(row['uid'])
            if row['sentByAgent']:
                ws1.append([row['uid'],row['email'],row['company'],'',dt])
            else:
                ws1.append([row['uid'],row['email'],row['company'],dt,''])

    # ws2 = wb.create_sheet(title="Pi")
    #
    # ws2['F5'] = 3.14
    #
    # ws3 = wb.create_sheet(title="Data")
    # for row in range(10, 20):
    #     for col in range(27, 54):
    #         _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
    # print(ws3['AA10'].value)
    # wb.save(filename = dest_filename)
    return wb


class ReviewFilterCalAPIView2(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        print '****** entered', self.request.GET ,' gggggggggggggg'
        # all chat threads
        toReturn = {}
        allUidInChat =  list(SupportChat.objects.all().values_list('uid',flat=True).distinct())
        chatThreadObj = ChatThread.objects.filter(uid__in=allUidInChat)
        # print chatThreadObj,'all chat thread obj'
        if 'customer' in self.request.GET:
            userCompany = list(service.objects.filter(contactPerson=self.request.user).values_list('pk',flat=True).distinct())
            userCustProfile = list(CustomerProfile.objects.filter(service__in=userCompany).values_list('pk',flat=True).distinct())
            chatThreadObj = chatThreadObj.filter(company__in=userCustProfile)
            if 'customerProfilePkList' in self.request.GET:
                return Response(userCustProfile, status=status.HTTP_200_OK)
            if 'agentComapnyPk' in self.request.GET:
                userCompany = list(service.objects.filter(advisors=self.request.user).values_list('pk',flat=True).distinct())
                userCustProfile = list(CustomerProfile.objects.filter(service__in=userCompany).values_list('pk',flat=True).distinct())
                return Response(userCustProfile, status=status.HTTP_200_OK)
        if 'getMyReviews' in self.request.GET:
            #when agent logged in
            chatThreadObj = chatThreadObj.filter(user = self.request.user)
        if 'date' in self.request.GET:
            date = datetime.datetime.strptime(self.request.GET['date'], '%Y-%m-%d').date()
            SupChatObj = SupportChat.objects.filter(created__startswith = date)
            uidL = list(SupChatObj.values_list('uid',flat=True).distinct())
            chatThreadObj = chatThreadObj.filter(uid__in=uidL)
        if 'user' in self.request.GET:
            user = User.objects.get(pk = int(self.request.GET['user']))
            chatThreadObj = chatThreadObj.filter(user = user)
        if 'client' in self.request.GET:
            userCustProfile = list(CustomerProfile.objects.filter(service=self.request.GET['client']).values_list('pk',flat=True).distinct())
            chatThreadObj = chatThreadObj.filter(company__in=userCustProfile)
        if 'email' in self.request.GET:
            uidList = list(Visitor.objects.filter(email=self.request.GET['email']).values_list('uid',flat=True).distinct())
            chatThreadObj = chatThreadObj.filter(uid__in=uidList)
        if 'audio' in self.request.GET:
            chatThreadObj = chatThreadObj.filter(typ = 'audio')
        if 'both' in self.request.GET:
            chatThreadObj = chatThreadObj.filter(Q(typ='audio')|Q(typ='video'))
        if 'video' in self.request.GET:
            chatThreadObj = chatThreadObj.filter(typ = 'video')
        if 'status' in self.request.GET:
            if self.request.GET['status']=='archived':
                chatThreadObj = chatThreadObj.filter(status = 'archived')
        else:
            chatThreadObj = chatThreadObj.filter(~Q(status = 'archived'))

        toReturn['dataLength'] = chatThreadObj.count()
        chatThreadList =  list(chatThreadObj.values())
        if 'sort' in self.request.GET:
            if self.request.GET['sortby']=='Created':
                chatThreadList = sorted(chatThreadList, key=lambda x: x['created'], reverse=True)
            if self.request.GET['sortby']=='Agent Name':
                chatThreadList = sorted(chatThreadList, key=lambda x: x['created'], reverse=True)
            if self.request.GET['sortby']=='UID':
                chatThreadList = sorted(chatThreadList, key=lambda x: x['uid'], reverse=True)
            if self.request.GET['sortby']=='Rating':
                chatThreadList = sorted(chatThreadList, key=lambda x: x['customerRating'], reverse=True)
            # print chatThreadList

        if 'limit' in self.request.GET and 'offset' in self.request.GET:
            print 'in limit'
            offset = int(self.request.GET['offset'])
            limit = offset+int(self.request.GET['limit'])
            chatThreadList = chatThreadList[offset:limit]
        toPush = []
        for idx,chatT in enumerate(chatThreadList):
            print chatT,"@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@"
            uidList = Visitor.objects.filter(uid=chatT['uid'])
            if len(uidList)>0:
                chatThreadList[idx]["Vname"] = uidList[0].name
            else:
                chatThreadList[idx]["Vname"] = False

        print len(chatThreadList)
        toReturn['data'] = chatThreadList
        # print toReturn,'toReturn'
        return Response(toReturn, status=status.HTTP_200_OK)


class ReviewHomeChatsAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        toReturn = []
        supChatObj = SupportChat.objects.all()
        if 'uid' in self.request.GET:
            supChatObj = supChatObj.filter(uid = self.request.GET['uid'])
        if 'date' in self.request.GET:
            date = datetime.datetime.strptime(self.request.GET['date'], '%Y-%m-%d').date()
            supChatObj = supChatObj.filter(created__startswith = date)
        toReturn = list(supChatObj.values())
        return Response(toReturn, status=status.HTTP_200_OK)


class ReviewFilterCalAPIView(APIView):
    renderer_classes = (JSONRenderer,)

    def get(self, request, format=None):
        print '****** entered', request.GET
        # offset=int(request.GET['offset'])
        # limit= offset+int(request.GET['limit'])
        toSend = []
        res = []
        sobj = SupportChat.objects.all()
        if 'customer' in self.request.GET:
            userCompany = list(service.objects.filter(contactPerson=self.request.user).values_list('pk',flat=True).distinct())
            userCustProfile = list(CustomerProfile.objects.filter(service__in=userCompany).values_list('pk',flat=True).distinct())
            if 'customerProfilePkList' in self.request.GET:
                print 'a################################################',userCustProfile
                return Response(userCustProfile, status=status.HTTP_200_OK)
            if 'agentComapnyPk' in self.request.GET:
                userCompany = list(service.objects.filter(advisors=self.request.user).values_list('pk',flat=True).distinct())
                userCustProfile = list(CustomerProfile.objects.filter(service__in=userCompany).values_list('pk',flat=True).distinct())

                return Response(userCustProfile, status=status.HTTP_200_OK)
            userCompanyUidList = list(ChatThread.objects.filter(company__in=userCustProfile).values_list('uid',flat=True).distinct())

            sobj = SupportChat.objects.filter(uid__in=userCompanyUidList)
        if 'getMyReviews' in self.request.GET:
            sobj = sobj.filter(user = self.request.user)
        if 'date' in self.request.GET:
            date = datetime.datetime.strptime(self.request.GET['date'], '%Y-%m-%d').date()
            sobj = sobj.filter(created__startswith = date)
        if 'user' in self.request.GET:
            sobj = sobj.filter(user = int(self.request.GET['user']))
        # if 'type' in self.request.GET:
        #     sobj = sobj.filter(user = int(self.request.GET['type']))
        if 'client' in self.request.GET:
            userCustProfile = list(CustomerProfile.objects.filter(service=self.request.GET['client']).values_list('pk',flat=True).distinct())
            userCompanyUidList = list(ChatThread.objects.filter(company__in=userCustProfile).values_list('uid',flat=True).distinct())
            # print userCustProfile,userCompanyUidList
            sobj = sobj.filter(uid__in=userCompanyUidList)
        # toSend = list(sobj.values())
        uidL = list(sobj.values_list('uid',flat=True).distinct())
        if 'status' in self.request.GET:
            if self.request.GET['status']=='archived':
                chObj1 = ChatThread.objects.filter(status='archived',uid__in=uidL)
                # agentsList = list(chObj).values_list('user',flat=True).distinct())
        else:
            chObj1 = ChatThread.objects.filter(~Q(status='archived'),uid__in=uidL)
        if 'audio' in self.request.GET:
            chObj1=chObj1.filter(Q(typ='audio'))
        if 'both' in self.request.GET:
            chObj1=chObj1.filter(Q(typ='audio')|Q(typ='video'))
        if 'video' in self.request.GET:
            chObj1=chObj1.filter(Q(typ='video'))

        agentsList=list(chObj1.values_list('user',flat=True).distinct())
        for i in agentsList:
            if 'status' in self.request.GET:
                if self.request.GET['status']=='archived':
                    chObj = ChatThread.objects.filter(status='archived',uid__in=uidL)
            else:
                chObj = ChatThread.objects.filter(~Q(status='archived'),user=i)
            if 'audio' in self.request.GET:
                chObj=chObj.filter(Q(typ='audio'))
            if 'both' in self.request.GET:
                chObj=chObj.filter(Q(typ='audio')|Q(typ='video'))
            if 'video' in self.request.GET:
                chObj=chObj.filter(Q(typ='video'))
            agentuidList=list(chObj.values_list('uid',flat=True).distinct())
            agSobj = sobj.filter(uid__in = agentuidList)
            if 'email' in self.request.GET:
                uidl = list(Visitor.objects.filter(email=self.request.GET['email']).values_list('uid',flat=True).distinct())
                agUid = list(agSobj.filter(uid__in=uidl).values_list('uid',flat=True).distinct())
            else:
                agUid = list(agSobj.values_list('uid',flat=True).distinct())

            # print agUid
            for j in agUid:
                cmntDate =  sobj.filter(uid = j)[0].created
                try:
                    email = Visitor.objects.get(uid=j).email
                except:
                    email = ''
                try:
                    company = ChatThread.objects.get(uid=j).company.service.name
                except:
                    company = ''
                try:
                    rating = ChatThread.objects.get(uid=j).customerRating
                except:
                    rating = ''
                try:
                    chatDuration = ChatThread.objects.get(uid=j).chatDuration
                except:
                    chatDuration = ''
                try:
                    numOfComments = ReviewComment.objects.filter(uid=j, chatedDate=cmntDate).count()
                except:
                    numOfComments = ''
                try:
                    statusChat = ChatThread.objects.get(uid=j).status
                except:
                    statusChat = ''
                try:
                    agentCommentCount = ReviewComment.objects.filter(uid=j, chatedDate=cmntDate, user = ChatThread.objects.get(uid=j).user).count()
                except:
                    agentCommentCount = ''
                try:
                    customerFeedback = ChatThread.objects.get(uid=j).customerFeedback
                except:
                    customerFeedback = ''
                try:
                    customerRating = ChatThread.objects.get(uid=j).customerRating
                except:
                    customerRating = ''
                try:
                    typ = ChatThread.objects.get(uid=j).typ
                except:
                    typ = ''
                try:
                    reviewedOn = ChatThread.objects.get(uid=j).reviewedOn
                except:
                    reviewedOn = ''
                try:
                    reviewedBy = ChatThread.objects.get(uid=j).reviewedBy
                except:
                    reviewedBy = ''
                try:
                    closedOn = ChatThread.objects.get(uid=j).closedOn
                except:
                    closedOn = ''
                try:
                    closedBy = ChatThread.objects.get(uid=j).closedBy
                except:
                    closedBy = ''
                try:
                    resolvedOn = ChatThread.objects.get(uid=j).resolvedOn
                except:
                    resolvedOn = ''
                try:
                    resolvedBy = ChatThread.objects.get(uid=j).resolvedBy
                except:
                    resolvedBy = ''
                try:
                    archivedOn = ChatThread.objects.get(uid=j).archivedOn
                except:
                    archivedOn = ''
                try:
                    archivedBy = ChatThread.objects.get(uid=j).archivedBy
                except:
                    archivedBy = ''
                try:
                    escalatedL1On = ChatThread.objects.get(uid=j).escalatedL1On
                except:
                    escalatedL1On = ''
                try:
                    escalatedL1By = ChatThread.objects.get(uid=j).escalatedL1By
                except:
                    escalatedL1By = ''
                try:
                    escalatedL2On = ChatThread.objects.get(uid=j).escalatedL2On
                except:
                    escalatedL2On = ''
                try:
                    escalatedL2By = ChatThread.objects.get(uid=j).escalatedL2By

                except:
                    escalatedL2By = ''
                # print agentCommentCount
                # print company
                agUidObj = list(agSobj.filter(uid=j).values().annotate(company=Value(company, output_field=CharField()) , rating=Value(rating, output_field=CharField()),
                chatDuration=Value(chatDuration, output_field=CharField()) , statusChat=Value(statusChat, output_field=CharField()) ,
                numOfComments=Value(numOfComments, output_field=CharField()),
                agentCommentCount=Value(agentCommentCount,output_field=CharField()), email=Value(email, output_field=CharField()),
                file=Concat(Value('/media/'),'attachment'),customerFeedback=Value(customerFeedback,output_field=CharField()),
                customerRating=Value(customerRating,output_field=CharField()),typ=Value(typ,output_field=CharField()),
                escalatedL2By=Value(escalatedL2By, output_field=CharField()),escalatedL2On=Value(escalatedL2On, output_field=CharField()),
                escalatedL1By=Value(escalatedL1By, output_field=CharField()),escalatedL1On=Value(escalatedL1On, output_field=CharField()),
                archivedBy=Value(archivedBy, output_field=CharField()),archivedOn=Value(archivedOn, output_field=CharField()),
                resolvedOn=Value(resolvedOn, output_field=CharField()),resolvedBy=Value(resolvedBy, output_field=CharField()),
                reviewedBy=Value(resolvedBy, output_field=CharField()),reviewedOn=Value(reviewedOn, output_field=CharField()),
                closedOn=Value(closedOn, output_field=CharField()),closedBy=Value(closedBy, output_field=CharField())))
                toSend.append(agUidObj)
                res = res + list(agSobj.filter(uid=j).values('uid','user','message','attachment','attachmentType','sentByAgent').annotate(company=Value(company, output_field=CharField()), rating=Value(rating, output_field=CharField()), numOfComments=Value(numOfComments, output_field=CharField()) , chatDuration=Value(chatDuration, output_field=CharField())  ,email=Value(email, output_field=CharField())))
        # print toSend
        if 'download' in self.request.GET:
            print 'downloadddddddddddddddddddddddd'
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=excelReport.xlsx'
            excelData = createExcel(res)
            print excelData,'dddddddddddddddd'
            excelData.save(response)
            return response
            # return ExcelResponse(res)
        toSend.reverse()
        return Response(toSend, status=status.HTTP_200_OK)


def encrypt(raw, password):
    private_key = hashlib.sha256(password.encode("utf-8")).digest()
    raw = pad(raw)
    iv = Random.new().read(AES.block_size)
    cipher = AES.new(private_key, AES.MODE_CBC, iv)
    return base64.b64encode(iv + cipher.encrypt(raw))


def decrypt(enc, password):
    private_key = hashlib.sha256(password.encode("utf-8")).digest()
    enc = base64.b64decode(enc)
    iv = enc[:16]
    cipher = AES.new(private_key, AES.MODE_CBC, iv)
    return unpad(cipher.decrypt(enc[16:]))

class GetChatterScriptAPI(APIView):
    def get(self , request , format = None):
        print '*******************************************'
        pk = request.GET['pk']
        encrypted = encrypt(pk, "cioc")
        print 'ee',encrypted

        while '/' in encrypted:
            encrypted = encrypt(pk, "cioc")
        print  decrypt(encrypted, "cioc")
        return Response({'data':encrypted  }, status = status.HTTP_200_OK)

class GetChatHistory(APIView):
    def get(self , request , format = None):
        if 'email' in request.GET:
            toSend = []
            email = request.GET['email']
            print email,'77777777777777777777'
            uidList = list(Visitor.objects.filter(email=email).values_list('uid',flat=True).distinct())
            print uidList
            for uid in uidList:
                chatList = list(SupportChat.objects.filter(uid=uid).values().annotate(file=Concat(Value('/media/'),'attachment')))
                agentList = list(SupportChat.objects.filter(uid=uid).values_list('user',flat=True).distinct())

                # print agentList,chatList
                toSend.append({'agentList':agentList,'chatList':chatList})
            return Response({'data':toSend}, status = status.HTTP_200_OK)
        else:
            return Response({}, status = status.HTTP_200_OK)


def getChatterScript(request , fileName):
    print fileName,'*****************'
    fileName = fileName.replace('.js' , '').replace("chatter-" , '')
    pk = decrypt(fileName , "cioc")
    print pk , 'HERRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRR'
    obj = CustomerProfile.objects.get(pk = pk)
    serviceWebsite = obj.service.web
    browserHeader =  dict((regex.sub('', header), value) for (header, value) in request.META.items() if header.startswith('HTTP_'))
    print browserHeader
    print request.META.get('HTTP_X_FORWARDED_FOR') , request.META.get('REMOTE_ADDR')
    print '**************8',browserHeader['REFERER'],serviceWebsite

    print globalSettings.SITE_ADDRESS
    print request.get_host()
    print request.META.get('REMOTE_ADDR')
    dataToSend = {"pk" : obj.pk ,'supportBubbleColor':obj.supportBubbleColor ,'iconColor':obj.iconColor, "windowColor" : obj.windowColor ,"fontColor":obj.fontColor,
        "custName" : obj.service.name , "chat":obj.chat , "callBack":obj.callBack , "video":obj.video ,"is_blink":obj.is_blink  ,
        "audio":obj.audio , "ticket":obj.ticket , "serverAddress" : globalSettings.SITE_ADDRESS ,
        "wampServer" : globalSettings.WAMP_SERVER ,"webrtcAddress": globalSettings.WEBRTC_ADDRESS,"wamp_prefix":globalSettings.WAMP_PREFIX,
        "chatIconPosition":obj.chatIconPosition,'chatIconType':obj.chatIconType}
    if obj.dp:
        dataToSend["dp"] =  obj.dp.url
    if obj.name:
        dataToSend["name"] =  obj.name
    if obj.support_icon:
        dataToSend["support_icon"] =  obj.support_icon.url
    else:
        dataToSend["support_icon"] = ''


    if obj.firstMessage:

        dataToSend["firstMessage"] =  obj.firstMessage
        print obj.firstMessage



    # return render(request, 'chatter.js', dataToSend ,content_type="application/x-javascript")
    if serviceWebsite in browserHeader['REFERER']:
        return render(request, 'chatter.js', dataToSend ,content_type="application/x-javascript")
    else:
        # pass
        return HttpResponse(request,'')
        # return render(request, 'chatter.js', dataToSend ,content_type="application/x-javascript")



class VisitorViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny,)
    serializer_class = VisitorSerializer
    queryset = Visitor.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['uid','email' ,'name','phoneNumber']

class ReviewCommentViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny,)
    serializer_class = ReviewCommentSerializer
    queryset = ReviewComment.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['uid','user','chatedDate']

class ChatThreadViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny,)
    serializer_class = ChatThreadSerializer
    queryset = ChatThread.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['uid','status','user','company']
    def get_queryset(self):
        if 'uid' in self.request.GET and 'checkThread' in self.request.GET:
            threadObj = ChatThread.objects.filter(uid = self.request.GET['uid'])
            if threadObj.count()>0:
                print 'sssssssssssssssssssss',threadObj[0].status
                if threadObj[0].status != 'started':
                    print 'nottttttttttttttt',threadObj[0].status
                    raise ValidationError(detail={'PARAMS' : 'createCookie'})
            print 'tttttttttttttttt',threadObj
            return threadObj
        return ChatThread.objects.all()
        if 'companyHandelrs' in self.request.GET and 'checkThread' in self.request.GET:
            threadObj = ChatThread.objects.filter(company = self.request.GET['companyHandelrs'])
            if threadObj.count()>0:
                print 'sssssssssssssssssssss',threadObj[0].status
                if threadObj[0].status != 'started':
                    print 'nottttttttttttttt',threadObj[0].status
                    raise ValidationError(detail={'PARAMS' : 'createCookie'})
            print 'tttttttttttttttt',threadObj
            return threadObj
        return ChatThread.objects.all()


class DocumentationViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny,)
    serializer_class = DocumentationSerializer
    queryset = Documentation.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['title','customer']

class GetChatTranscriptsViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny,)
    serializer_class = ChatThreadSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['uid','status']
    def get_queryset(self):
        if 'apiKey' in self.request.GET:
            try:
                userExit = CustomerProfile.objects.get(userApiKey=self.request.GET['apiKey'])
            except:
                raise PermissionDenied()
            if 'date' in self.request.GET:
                try:
                    date = datetime.datetime.strptime(self.request.GET['date'], '%Y-%m-%d').date()
                except:
                    raise ValidationError(detail={'PARAMS' : 'Date Argument Should Be yyyy-mm-dd Formate'} )
                print date,'dateeeeeeeeeeee'
                return ChatThread.objects.filter(created__startswith = date)
            else:
                raise ValidationError(detail={'PARAMS' : 'Date Argument Is Required'} )
        else:
            raise PermissionDenied()

class GetVisitorDetailsViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny,)
    serializer_class = VisitorSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['uid','email' ,'name']
    def get_queryset(self):
        if 'apiKey' in self.request.GET:
            try:
                userExit = CustomerProfile.objects.get(userApiKey=self.request.GET['apiKey'])
            except:
                raise PermissionDenied()
            if 'date' in self.request.GET:
                try:
                    date = datetime.datetime.strptime(self.request.GET['date'], '%Y-%m-%d').date()
                except:
                    raise ValidationError(detail={'PARAMS' : 'Date Argument Should Be yyyy-mm-dd Formate'} )
                print date,'dateeeeeeeeeeee'
                return Visitor.objects.filter(created__startswith = date)
            else:
                raise ValidationError(detail={'PARAMS' : 'Date Argument Is Required'} )
        else:
            raise PermissionDenied()

class GetOfflineMessagesViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny,)
    serializer_class = SupportChatSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['uid','user']
    exclude_fields = ['id']
    def get_queryset(self):
        if 'apiKey' in self.request.GET:
            try:
                userExit = CustomerProfile.objects.get(userApiKey=self.request.GET['apiKey'])
            except:
                raise PermissionDenied()
            if 'date' in self.request.GET:
                try:
                    date = datetime.datetime.strptime(self.request.GET['date'], '%Y-%m-%d').date()
                except:
                    raise ValidationError(detail={'PARAMS' : 'Date Argument Should Be yyyy-mm-dd Formate'} )
                print date,'dateeeeeeeeeeee'
                return SupportChat.objects.filter(user__isnull=True,created__startswith = date)
            else:
                raise ValidationError(detail={'PARAMS' : 'Date Argument Is Required'} )
        else:
            raise PermissionDenied()

class GethomeCal2(APIView):
    def get(self , request , format = None):
        today = datetime.datetime.now().date()
        tomorrow = today + relativedelta(days=1)
        lastWeek = today - relativedelta(days=6)
        lastToLastWeek =  lastWeek - relativedelta(days=7)
        chatThreadObj = ChatThread.objects.filter(created__range=(lastWeek,tomorrow))
        chatThreadObjOld = ChatThread.objects.filter(created__range=(lastToLastWeek,lastWeek - relativedelta(days=1)))

        changeInData = {}
        changeInChat = {'percentage':0 , 'increase' : False}
        changeInAvgChatDur = {'percentage':0 , 'increase' : False}
        changeInFrtAvg = {'percentage':0 , 'increase' : False}
        changeInRespTimeAvg = {'percentage':0 , 'increase' : False}
        changeInMissedChat = {'percentage':0 , 'increase' : False}
        changeInAverageRating = {'percentage':0 , 'increase' : False}

        if 'client' in self.request.GET:
            if int(self.request.GET['company'])>0:
                print 'clientclientclient',int(self.request.GET['company'])
                chatThreadObj = chatThreadObj.filter(company=int(self.request.GET['company']))
                print chatThreadObj,'chatttt'
        else:
            print 'not client'

        if 'agent' in self.request.GET:
            if int(self.request.GET['company'])>0:
                print 'agentagentagent',int(self.request.GET['company'])
                user = User.objects.get(pk = int(self.request.GET['user']))
                chatThreadObj = chatThreadObj.filter(company=int(self.request.GET['company']), user = user)
                print chatThreadObj
        else:
            print 'not agent'

        totalChatCount = chatThreadObj.count()
        totalAudioCalls = chatThreadObj.filter(typ = 'audio').count()
        totalVideoCalls = chatThreadObj.filter(typ = 'video').count()
        totalChatMessage = chatThreadObj.filter(~Q(typ='audio') & ~Q(typ='video')).count()
        totalChatCountOld = chatThreadObjOld.count()
        missedChatCount = chatThreadObj.filter(user__isnull=True).count()
        missedChatCountOld = chatThreadObjOld.filter(user__isnull=True).count()

        arAll = chatThreadObj.filter(customerRating__isnull=False).aggregate(Avg('customerRating'))
        avgRatingAll = arAll['customerRating__avg'] if arAll['customerRating__avg'] else 0

        arAllOld = chatThreadObjOld.filter(customerRating__isnull=False).aggregate(Avg('customerRating'))
        avgRatingAllOld = arAllOld['customerRating__avg'] if arAllOld['customerRating__avg'] else 0


        frtAll = chatThreadObj.filter(firstResponseTime__isnull=False).aggregate(Avg('firstResponseTime'))
        firstResTimeAvgAll = frtAll['firstResponseTime__avg'] if frtAll['firstResponseTime__avg'] else 0

        frtAllOld  = chatThreadObjOld.filter(firstResponseTime__isnull=False).aggregate(Avg('firstResponseTime'))
        firstResTimeAvgAllOld = frtAllOld['firstResponseTime__avg'] if frtAllOld['firstResponseTime__avg'] else 0

        uidL = list(chatThreadObj.values_list('uid', flat = True).distinct())
        supportChats = SupportChat.objects.filter(uid__in = uidL)

        uidLOld = list(chatThreadObjOld.values_list('uid', flat = True).distinct())
        supportChatsOld = SupportChat.objects.filter(uid__in = uidLOld)

        avgResTAll = supportChats.filter(responseTime__isnull=False).aggregate(Avg('responseTime'))
        avgRespTimeAll = avgResTAll['responseTime__avg'] if avgResTAll['responseTime__avg'] else 0

        avgResTOldAll = supportChatsOld.filter(responseTime__isnull=False).aggregate(Avg('responseTime'))
        avgRespTimeAllOld = avgResTOldAll['responseTime__avg'] if avgResTOldAll['responseTime__avg'] else 0

        achatDur = chatThreadObj.filter(chatDuration__isnull=False).aggregate(Avg('chatDuration'))
        avgChatDuration = achatDur['chatDuration__avg'] if achatDur['chatDuration__avg'] else 0

        achatDurOld = chatThreadObjOld.filter(chatDuration__isnull=False).aggregate(Avg('chatDuration'))
        avgChatDurationOld = achatDurOld['chatDuration__avg'] if achatDurOld['chatDuration__avg'] else 0

        if totalChatCountOld<totalChatCount:
            changeInChat['percentage'] = (float(totalChatCount - totalChatCountOld)/totalChatCount)*100
            changeInChat['increase'] = True
        elif totalChatCount<totalChatCountOld:
            changeInChat['percentage'] = (float(totalChatCountOld - totalChatCount)/totalChatCountOld)*100
            changeInChat['increase'] = False
        else:
            changeInChat['percentage'] = 0.0
            changeInChat['increase'] = False

        if missedChatCountOld<missedChatCount:
            changeInMissedChat['percentage'] = (float(missedChatCount - missedChatCountOld)/missedChatCount)*100
            changeInMissedChat['increase'] = True
        elif missedChatCount< missedChatCountOld:
            changeInMissedChat['percentage'] = (float(missedChatCountOld - missedChatCount)/missedChatCountOld)*100
            changeInMissedChat['increase'] = False
        else:
            changeInMissedChat['percentage'] = 0.0
            changeInMissedChat['increase'] = False

        print avgRatingAllOld, avgRatingAll ,'avgRatingAllOld'
        if avgRatingAllOld<avgRatingAll:
            changeInAverageRating['percentage'] = (float(avgRatingAll - avgRatingAllOld)/avgRatingAll)*100
            changeInAverageRating['increase'] = True
        elif avgRatingAll<avgRatingAllOld:
            changeInAverageRating['percentage'] = (float(avgRatingAllOld - avgRatingAll)/avgRatingAllOld)*100
            changeInAverageRating['increase'] = False
        else:
            changeInAverageRating['percentage'] = 0.0
            changeInAverageRating['increase'] = False

        if firstResTimeAvgAllOld<firstResTimeAvgAll:
            changeInFrtAvg['percentage'] = (float(firstResTimeAvgAll - firstResTimeAvgAllOld)/firstResTimeAvgAll)*100
            changeInFrtAvg['increase'] = True
        elif firstResTimeAvgAll<firstResTimeAvgAllOld:
            changeInFrtAvg['percentage'] = (float(firstResTimeAvgAllOld - firstResTimeAvgAll)/firstResTimeAvgAllOld)*100
            changeInFrtAvg['increase'] = False
        else:
            changeInFrtAvg['percentage'] = 0.0
            changeInFrtAvg['increase'] = False

        if avgRespTimeAllOld<avgRespTimeAll:
            changeInRespTimeAvg['percentage'] = (float(avgRespTimeAll - avgRespTimeAllOld)/avgRespTimeAll)*100
            changeInRespTimeAvg['increase'] = True
        elif avgRespTimeAll<avgRespTimeAllOld:
            changeInRespTimeAvg['percentage'] = (float(avgRespTimeAllOld - avgRespTimeAll)/avgRespTimeAllOld)*100
            changeInRespTimeAvg['increase'] = False
        else:
            changeInRespTimeAvg['percentage'] = 0.0
            changeInRespTimeAvg['increase'] = False

        if avgChatDurationOld<avgChatDuration:
            changeInAvgChatDur['percentage'] = (float(avgChatDuration - avgChatDurationOld)/avgChatDuration)*100
            changeInAvgChatDur['increase'] = True
        elif avgChatDuration<avgChatDurationOld:
            changeInAvgChatDur['percentage'] = (float(avgChatDurationOld - avgChatDuration)/avgChatDurationOld)*100
            changeInAvgChatDur['increase'] = False
        else:
            changeInAvgChatDur['percentage'] = 0.0
            changeInAvgChatDur['increase'] = False

        changeInData['changeInChat'] = changeInChat
        changeInData['changeInMissedChat'] = changeInMissedChat
        changeInData['changeInAverageRating'] = changeInAverageRating
        changeInData['changeInFrtAvg'] = changeInFrtAvg
        changeInData['changeInRespTimeAvg'] = changeInRespTimeAvg
        changeInData['changeInAvgChatDur'] =changeInAvgChatDur


        agentLeaderBoard = []
        agL = list(chatThreadObj.filter(user__isnull=False).values_list('user',flat=True).distinct())
        print agL,'aglllllllllll'
        for i in agL:
            oneAgentChatThreadObj = chatThreadObj.filter(user=i)
            rA = oneAgentChatThreadObj.filter(customerRating__isnull=False).aggregate(Avg('customerRating'))
            ratingAvg = rA['customerRating__avg'] if rA['customerRating__avg'] else 0

            uidL = list(oneAgentChatThreadObj.values_list('uid', flat = True).distinct())
            supportChats = SupportChat.objects.filter(uid__in = uidL)


            rtA = supportChats.filter(responseTime__isnull=False).aggregate(Avg('responseTime'))
            respTimeAvg = rtA['responseTime__avg'] if rtA['responseTime__avg'] else 0

            frtA = oneAgentChatThreadObj.filter(firstResponseTime__isnull=False).aggregate(Avg('firstResponseTime'))
            firstResTimeAvg = frtA['firstResponseTime__avg'] if frtA['firstResponseTime__avg'] else 0

            noOfChats = oneAgentChatThreadObj.count()

            cdA = oneAgentChatThreadObj.filter(chatDuration__isnull=False).aggregate(Avg('chatDuration'))
            chatDurationAvg = cdA['chatDuration__avg'] if cdA['chatDuration__avg'] else 0
            toAppend = {'agentPk':i,'ratingAvg':ratingAvg ,'respTimeAvg':respTimeAvg ,'firstResTimeAvg':firstResTimeAvg,'noOfChats':noOfChats,'chatDurationAvg':chatDurationAvg}
            print toAppend,'ABCD'
            agentLeaderBoard.append(toAppend)


        graphData = [[],[]]
        graphLabels = []
        for i in range(7):
            dt = lastWeek + relativedelta(days=i)
            dateChat = chatThreadObj.filter(created__startswith=dt)
            if 'perticularUser' in self.request.GET:
                if int(self.request.GET['perticularUser'])>0:
                    dateChat = dateChat.filter(company=int(self.request.GET['perticularUser']))
            missed = dateChat.filter(user__isnull=True).count()
            received = dateChat.filter(user__isnull=False).count()
            graphData[0].append(received)
            graphData[1].append(missed)
            graphLabels.append(datetime.datetime.combine(dt, datetime.datetime.min.time()).strftime('%b %d'))
        toSend = {'totalChatCount':totalChatCount,'totalChatMessage':totalChatMessage,'totalVideoCalls':totalVideoCalls,'totalAudioCalls':totalAudioCalls,'missedChatCount':missedChatCount,'changeInData':changeInData,'avgRatingAll':avgRatingAll,'avgRespTimeAll':avgRespTimeAll,'firstResTimeAvgAll':firstResTimeAvgAll,'avgChatDuration':avgChatDuration,'graphData':graphData,'graphLabels':graphLabels,'agentLeaderBoard':agentLeaderBoard}
        # print toSend ,'toSendtoSendtoSendtoSendtoSend'
        print type(toSend)
        return Response(toSend, status = status.HTTP_200_OK)

class DocumentVersionViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny,)
    serializer_class = DocumentVersionSerializer
    queryset = DocumentVersion.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['parent']


class CompanyProcessViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny,)
    serializer_class = CompanyProcessSerializer
    # queryset = CompanyProcess.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['service','text']
    def get_queryset(self):
        print '@@@@@@@@@@@@@@@@@@@@@@@'
        return CompanyProcess.objects.all()

class CannedResponsesViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny,)
    serializer_class = CannedResponsesSerializer
    queryset = CannedResponses.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['text','service']

class DynamicFormViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny,)
    serializer_class = DynamicFormSerializer
    # queryset = DynamicForm.objects.all()
    def get_queryset(self):
        if 'companyPk' in  self.request.GET:
            return DynamicForm.objects.filter(company = self.request.GET['companyPk'])
        return DynamicForm.objects.all()

class DynamicFieldViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny,)
    serializer_class = DynamicFieldSerializer
    queryset = DynamicField.objects.all()
    def get_queryset(self):
        if 'formPk' in  self.request.GET:
            return DynamicField.objects.filter(form = self.request.GET['formPk'])
        return DynamicField.objects.all()




class HeartbeatApi(APIView):
    permission_classes = (permissions.AllowAny,)
    def get(self , request , format = None):
        if 'timesheet' in request.GET:
            u = User.objects.get(pk = request.GET['pk'])
            today_min = datetime.datetime.combine(datetime.date.today(), datetime.time.min)
            today_max = datetime.datetime.combine(datetime.date.today(), datetime.time.max)
            obj=Heartbeat.objects.filter(start__range=(today_min, today_max),user=u)
            print obj ,"heartbeat objjjjjjjjjj"
            if len(obj)==1:
                if obj[0].end is None:
                    obj[0].end=timezone.now()
                    difference=obj[0].end-obj[0].start
                    obj[0].duration = difference.total_seconds()
                    day_difference=difference
                    obj[0].day_duration = day_difference.total_seconds()
                    obj[0].save()
                else:
                    c = timezone.now() - obj[0].end
                    cInMin =  c.total_seconds()/60
                    if cInMin>15:
                        Heartbeat.objects.create(start=timezone.now(),user=u)
                    else:
                        obj[0].end=timezone.now()
                        difference=obj[0].end-obj[0].start
                        obj[0].duration = difference.total_seconds()
                        day_difference=difference
                        obj[0].day_duration = day_difference.total_seconds()
                        obj[0].save()
            elif len(obj)>1:
                print 'multiple hearbeat'
                if obj[len(obj)-1].end is None:
                    print 'heartbeat has only start'
                    obj[len(obj)-1].end=timezone.now()
                    difference=obj[len(obj)-1].end-obj[len(obj)-1].start
                    obj[len(obj)-1].duration = difference.total_seconds()
                    day_difference=obj[len(obj)-1].end-obj[0].start
                    obj[0].day_duration = day_difference.total_seconds()
                    obj[0].save()
                    obj[len(obj)-1].save()
                else:
                    print 'heartbeat has end'
                    c = timezone.now() - obj[len(obj)-1].end
                    cInMin =  c.total_seconds()/60
                    if cInMin>15:
                        print 'time exceeds 30 mints creating new heartbeat'
                        Heartbeat.objects.create(start=timezone.now(),user=u)
                    else:
                        print 'updating end for the heartbeat'
                        obj[len(obj)-1].end=timezone.now()
                        difference=obj[len(obj)-1].end-obj[len(obj)-1].start
                        day_difference=obj[len(obj)-1].end-obj[0].start
                        obj[0].day_duration = day_difference.total_seconds()
                        obj[len(obj)-1].duration = difference.total_seconds()
                        obj[0].save()
                        obj[len(obj)-1].save()
            else:
                Heartbeat.objects.create(start=timezone.now(),user=u)
            return Response({}, status = status.HTTP_200_OK)
        elif 'getDetailData' in request.GET:
            u = User.objects.get(pk = request.GET['pk'])
            # sobj = sobj.filter(user = self.request.user)
            toSend=[]
            heartbtObj = Heartbeat.objects.all()
            if 'date' in request.GET:
                date = datetime.datetime.strptime(request.GET['date'], '%Y-%m-%d').date()
                o = heartbtObj.filter(created__startswith = date,user=u)
                th=list(o.values())
                toSend.append(th);
                return Response(toSend, status=status.HTTP_200_OK)
            else:
                dates = list(heartbtObj.filter(user=u).values_list('created',flat=True).distinct())
                for j in dates:
                    o=heartbtObj.filter(created__startswith = j,user=u)
                    th=list(o.values())
                    toSend.append(th);
                return Response(toSend, status=status.HTTP_200_OK)
        elif 'getTimeSheetData' in request.GET:
            newDetails=[]
            heartbtObj = Heartbeat.objects.filter(created__startswith=request.GET['date'])
            uidL = list(heartbtObj.values_list('user',flat=True).distinct())
            for j in uidL:
                u = User.objects.get(pk = j)
                o = heartbtObj.filter(user=u)
                th=list(o.values())
                newDetails.append(th);
            return Response(newDetails, status=status.HTTP_200_OK)
class StreamRecordings(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes=(permissions.AllowAny,)
    def post(self , request , format = None):
        print request.data,'dddddddddddddddd'
        filename = request.data['fileName']
        filepath=os.path.join(globalSettings.BASE_DIR,'media_root',filename)
        with open(filepath, 'wb+') as destination:
            for chunk in request.FILES['file'].chunks():
                destination.write(chunk)
        return Response({}, status = status.HTTP_200_OK)


class getChatStatus(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes=(permissions.AllowAny,)
    def get(self , request , format = None):
        print request.data,'dddddddddddddddd'
        uid=request.GET['uid']
        sendMail=False;
        changeStatus=False;
        chatT= ChatThread.objects.filter(uid=uid)[0]
        compPk=chatT.company.pk
        if 'sendMail' in request.GET:
            diff=timezone.now()-chatT.lastActivity
            diffInMin = diff.total_seconds()/60
            if diffInMin>10 and chatT.status=='started' and chatT.mailSent is None:
                sendMail=True;
                chatT.mailSent=timezone.now();
                chatT.save()
            elif diffInMin>10 and chatT.status=='started' and chatT.lastActivity>chatT.mailSent:
                sendMail=True;
                chatT.mailSent=timezone.now();
                chatT.save()
        if 'checkStatus' in request.GET:
            diffForstatus=timezone.now()-chatT.created
            diffForstatusInMints=diffForstatus.total_seconds()/60
            if diffForstatusInMints>10 and chatT.user is None:
                changeStatus=True;
                chatT.status="missed"
                chatT.save()
            if diffForstatusInMints>3 and chatT.user is None:
                chatT.isLate=True
                chatT.save()
        return Response({'sendMail':sendMail,'changeStatus':changeStatus,'companyPk':compPk}, status = status.HTTP_200_OK)


class CreateSVG(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes=(permissions.AllowAny,)
    def post(self , request , format = None):
        nameOfFile="demo"
        color='black';
        strokeWidth='2';
        fill="white";
        path=request.data['path'];
        svgHeight=500;
        svgWidth=500;

        if 'strokeWidth' in request.data:
            strokeWidth=request.data['strokeWidth'];
        if 'fill' in request.data:
            fill=request.data['fill'];
        if 'color' in request.data:
            color=request.data['color'];
        if 'fileName' in request.data:
            nameOfFile=request.data['fileName'];
        if 'svgHeight' in request.data:
            svgHeight=request.data['svgHeight'];
        if 'svgWidth' in request.data:
            svgWidth=request.data['svgWidth'];

        filepath=os.path.join(globalSettings.BASE_DIR,'media_root','index.svg')
        filepathOfDest=os.path.join(globalSettings.BASE_DIR,'media_root',nameOfFile+'.svg')


        # text_to_search='<svg></svg>'
        replacement_text="<path d='"+path+"' stroke='"+color+"' stroke-width='"+strokeWidth+"' fill='"+fill+"'/>"
        # replacement_text='<svg height="'+svgHeight+'" width="'+svgWidth+'"><path d="'+path+'" stroke="'+color+'" stroke-width="'+strokeWidth+'" fill="'+fill+'"/></svg>'


        # with open(filepath) as f:
        #     newText=f.read().replace(text_to_search, replacement_text)
        # with open(filepathOfDest, "w") as f:
        #     f.write(newText)
        print replacement_text,"***************8"
        return Response({'path':replacement_text}, status = status.HTTP_200_OK)


class EmailChat(APIView):
    renderer_classes = (JSONRenderer,)
    def post(self , request , format = None):
        emailAddr=[]
        print request.data['email'],'email'
        emailAddr.append(request.data['email'])
        chatStarted=ChatThread.objects.filter(uid=request.data['uid'])[0].created
        sObj = SupportChat.objects.filter(uid = request.data['uid'])
        visitor = Visitor.objects.filter(uid = request.data['uid'])

        if len(visitor)>0:
            name = visitor[0].name
        else:
            name=request.data['uid']
        allChats = []
        for a in sObj:
            print a.created,"*******************************"
            toAppend = {'user': a.user , 'message': a.message , 'created': a.created , 'uid': name }
            if a.sentByAgent:
                toAppend['sentByAgent'] = True
                if a.attachment:
                    toAppend['attachment'] = globalSettings.SITE_ADDRESS + '/media/' + str(a.attachment)
            else:
                toAppend['sentByAgent'] = False
                if a.attachment:
                    toAppend['attachment'] = globalSettings.SITE_ADDRESS + '/media/' + str(a.attachment)
            allChats.append(toAppend)
        sObj = allChats
        ctx = {
            'heading' : "Support Conversation",
            'allChats' : sObj,
            'started':chatStarted
        }
        print ctx
        email_body = get_template('app.support.email.html').render(ctx)
        msg = EmailMessage("Chat Conversation" , email_body, to= emailAddr)
        msg.content_subtype = 'html'
        msg.send()
        return Response({}, status = status.HTTP_200_OK)




class EmailScript(APIView):
    renderer_classes = (JSONRenderer,)
    def post(self , request , format = None):
        emailAddr=[]
        print request.data['email'],'email'
        emailAddr.append(request.data['email'])
        sObj = request.data['script']
        companyName=request.data['companyName']


        ctx = {
            'heading' : "Syrow Script",
            'script' : sObj,
            'companyName':companyName

        }
        print ctx
        email_body = get_template('app.scriptEmail.html').render(ctx)
        msg = EmailMessage("Syrow chat support installation guide" , email_body, to= emailAddr)
        msg.content_subtype = 'html'
        msg.send()
        return Response({}, status = status.HTTP_200_OK)
